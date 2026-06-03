import pandas as pd
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Microsoft YaHei", size=9)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14)
SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=11)

P1_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
P2_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
P3_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISSING_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_report(mapping_df, diff_df, hk_df, dahua_df, output_path):
    print(f"[report] Generating Excel report: {output_path}")
    wb = Workbook()

    _write_mapping_overview(wb, mapping_df)
    _write_param_diff_detail(wb, diff_df)
    _write_series_coverage(wb, hk_df, dahua_df)
    _write_data_quality(wb, diff_df)
    _write_run_summary(wb, mapping_df, diff_df, hk_df, dahua_df)

    wb.save(output_path)
    print(f"[report] Report saved to {output_path}")


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_cell(cell, value=""):
    cell.font = DATA_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER


def _auto_width(ws, max_col, min_w=10, max_w=40):
    for col in range(1, max_col + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2


def _write_mapping_overview(wb, mapping_df):
    ws = wb.active
    ws.title = "Mapping Overview"

    ws.cell(row=1, column=1, value="Cross-Brand Model Mapping Overview").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = [
        "Mapping ID", "Confidence", "Match Dimensions",
        "HK Series", "HK Subseries", "HK Model", "HK Name",
        "Dahua Series", "Dahua Subseries", "Dahua Model", "Dahua Name",
        "Diff Count",
    ]

    row = 3
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    _style_header(ws, row, len(headers))

    if mapping_df.empty:
        ws.cell(row=4, column=1, value="No mappings found")
        return

    diff_count_map = {}
    if diff_df is not None and not diff_df.empty:
        non_match = diff_df[diff_df["diff_type"].isin(["value_diff", "feature_diff", "one_side_missing"])]
        diff_count_map = non_match.groupby("mapping_id").size().to_dict()

    for idx, (_, m) in enumerate(mapping_df.iterrows()):
        r = row + 1 + idx
        mid = m.get("mapping_id", "")
        values = [
            mid,
            m.get("confidence_score", 0),
            m.get("match_dimensions", ""),
            m.get("hikvision_series_l1", ""),
            m.get("hikvision_subseries", ""),
            m.get("hikvision_model", ""),
            m.get("hikvision_name", ""),
            m.get("dahua_series_l1", ""),
            m.get("dahua_subseries", ""),
            m.get("dahua_model", ""),
            m.get("dahua_name", ""),
            diff_count_map.get(mid, 0),
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            _style_data_cell(cell)

        score = m.get("confidence_score", 0)
        score_cell = ws.cell(row=r, column=2)
        if score >= 0.8:
            score_cell.fill = MATCH_FILL
        elif score >= 0.6:
            score_cell.fill = LIGHT_BLUE_FILL

    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row + len(mapping_df)}"
    _auto_width(ws, len(headers))
    ws.freeze_panes = f"A{row + 1}"


def _write_param_diff_detail(wb, diff_df):
    ws = wb.create_sheet("Param Diff Detail")

    ws.cell(row=1, column=1, value="Parameter Difference Detail").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    headers = [
        "Mapping ID", "HK Model", "Dahua Model",
        "Field Code", "HK Field Name", "Dahua Field Name",
        "HK Value", "Dahua Value",
        "Diff Type", "Diff Detail", "Severity",
    ]

    row = 3
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    _style_header(ws, row, len(headers))

    if diff_df is None or diff_df.empty:
        ws.cell(row=4, column=1, value="No diff records")
        return

    for idx, (_, d) in enumerate(diff_df.iterrows()):
        r = row + 1 + idx
        values = [
            d.get("mapping_id", ""),
            d.get("hikvision_model", ""),
            d.get("dahua_model", ""),
            d.get("field_code", ""),
            d.get("field_name_hk", ""),
            d.get("field_name_dahua", ""),
            d.get("hikvision_value", ""),
            d.get("dahua_value", ""),
            d.get("diff_type", ""),
            d.get("diff_detail", ""),
            d.get("severity", ""),
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            _style_data_cell(cell)

        severity = d.get("severity", "")
        sev_cell = ws.cell(row=r, column=11)
        if severity == "P1":
            sev_cell.fill = P1_FILL
            sev_cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=9)
        elif severity == "P2":
            sev_cell.fill = P2_FILL
        elif severity == "P3":
            sev_cell.fill = P3_FILL

        diff_type = d.get("diff_type", "")
        dt_cell = ws.cell(row=r, column=9)
        if diff_type == "match":
            dt_cell.fill = MATCH_FILL
        elif diff_type == "one_side_missing":
            dt_cell.fill = MISSING_FILL

    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row + len(diff_df)}"
    _auto_width(ws, len(headers))
    ws.freeze_panes = f"A{row + 1}"


def _write_series_coverage(wb, hk_df, dahua_df):
    ws = wb.create_sheet("Series Coverage")

    ws.cell(row=1, column=1, value="Series Coverage Comparison").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ["Dimension", "Hikvision", "Dahua", "Difference"]
    row = 3
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    _style_header(ws, row, len(headers))

    rows_data = []

    hk_series = hk_df["series_l1"].nunique() if "series_l1" in hk_df.columns else 0
    dahua_series = dahua_df["series_l1"].nunique() if "series_l1" in dahua_df.columns else 0
    rows_data.append(("L1 Series Count", hk_series, dahua_series, hk_series - dahua_series))

    hk_sub = hk_df["series_l2"].nunique() if "series_l2" in hk_df.columns else 0
    dahua_sub = dahua_df["series_l2"].nunique() if "series_l2" in dahua_df.columns else 0
    rows_data.append(("L2 Subseries Count", hk_sub, dahua_sub, hk_sub - dahua_sub))

    rows_data.append(("Total Models", len(hk_df), len(dahua_df), len(hk_df) - len(dahua_df)))

    if "series_l1" in hk_df.columns:
        for s in sorted(hk_df["series_l1"].unique()):
            count = len(hk_df[hk_df["series_l1"] == s])
            rows_data.append((f"HK - {s}", count, "-", "-"))

    if "series_l1" in dahua_df.columns:
        for s in sorted(dahua_df["series_l1"].unique()):
            count = len(dahua_df[dahua_df["series_l1"] == s])
            rows_data.append((f"Dahua - {s}", "-", count, "-"))

    for i, (dim, hk_val, dahua_val, diff_val) in enumerate(rows_data):
        r = row + 1 + i
        for col_idx, v in enumerate([dim, hk_val, dahua_val, diff_val], 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            _style_data_cell(cell)

    _auto_width(ws, len(headers))


def _write_data_quality(wb, diff_df):
    ws = wb.create_sheet("Data Quality")

    ws.cell(row=1, column=1, value="Data Quality Summary").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    if diff_df is None or diff_df.empty:
        ws.cell(row=3, column=1, value="No diff data available")
        return

    headers = ["Diff Type", "Count", "Percentage", "Top Fields"]
    row = 3
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=row, column=col_idx, value=h)
    _style_header(ws, row, len(headers))

    type_counts = diff_df["diff_type"].value_counts()
    total = len(diff_df)

    for i, (dtype, count) in enumerate(type_counts.items()):
        r = row + 1 + i
        pct = f"{count / total * 100:.1f}%"
        subset = diff_df[diff_df["diff_type"] == dtype]
        top_fields = subset["field_code"].value_counts().head(3)
        top_str = ", ".join([f"{f}({c})" for f, c in top_fields.items()])

        for col_idx, v in enumerate([dtype, count, pct, top_str], 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            _style_data_cell(cell)

    r = row + 2 + len(type_counts)
    ws.cell(row=r, column=1, value="Severity Distribution").font = SUBTITLE_FONT
    r += 1

    sev_headers = ["Severity", "Count", "Percentage"]
    for col_idx, h in enumerate(sev_headers, 1):
        ws.cell(row=r, column=col_idx, value=h)
    _style_header(ws, r, len(sev_headers))

    sev_counts = diff_df[diff_df["severity"] != "-"]["severity"].value_counts()
    sev_total = sev_counts.sum()
    for i, (sev, count) in enumerate(sev_counts.items()):
        r2 = r + 1 + i
        pct = f"{count / sev_total * 100:.1f}%" if sev_total > 0 else "0%"
        for col_idx, v in enumerate([sev, count, pct], 1):
            cell = ws.cell(row=r2, column=col_idx, value=v)
            _style_data_cell(cell)
            if sev == "P1" and col_idx == 1:
                cell.fill = P1_FILL
            elif sev == "P2" and col_idx == 1:
                cell.fill = P2_FILL
            elif sev == "P3" and col_idx == 1:
                cell.fill = P3_FILL

    _auto_width(ws, max(len(headers), len(sev_headers)))


def _write_run_summary(wb, mapping_df, diff_df, hk_df, dahua_df):
    ws = wb.create_sheet("Run Summary")

    ws.cell(row=1, column=1, value="Analysis Run Summary").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    items = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Hikvision Models", len(hk_df)),
        ("Dahua Models", len(dahua_df)),
        ("Total Mappings", len(mapping_df) if mapping_df is not None else 0),
        ("Total Diff Records", len(diff_df) if diff_df is not None and not diff_df.empty else 0),
    ]

    if mapping_df is not None and not mapping_df.empty:
        items.extend([
            ("Avg Confidence", f"{mapping_df['confidence_score'].mean():.3f}"),
            ("Max Confidence", f"{mapping_df['confidence_score'].max():.3f}"),
            ("Min Confidence", f"{mapping_df['confidence_score'].min():.3f}"),
            ("High Confidence (>=0.8)", len(mapping_df[mapping_df["confidence_score"] >= 0.8])),
            ("Medium Confidence (0.6-0.8)", len(mapping_df[
                (mapping_df["confidence_score"] >= 0.6) & (mapping_df["confidence_score"] < 0.8)
            ])),
            ("Low Confidence (<0.6)", len(mapping_df[mapping_df["confidence_score"] < 0.6])),
        ])

    if diff_df is not None and not diff_df.empty:
        type_counts = diff_df["diff_type"].value_counts()
        for dtype, count in type_counts.items():
            items.append((f"Diff: {dtype}", count))

    row = 3
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate competitor analysis Excel report")
    parser.add_argument("--mapping-csv", required=True)
    parser.add_argument("--diff-csv", required=True)
    parser.add_argument("--hk-csv", required=True)
    parser.add_argument("--dahua-csv", required=True)
    parser.add_argument("--output", required=True, help="Output Excel file path")
    args = parser.parse_args()

    mapping_df = pd.read_csv(args.mapping_csv, encoding="utf-8-sig")
    diff_df = pd.read_csv(args.diff_csv, encoding="utf-8-sig")
    hk_df = pd.read_csv(args.hk_csv, low_memory=False)
    dahua_df = pd.read_csv(args.dahua_csv, low_memory=False)
    generate_report(mapping_df, diff_df, hk_df, dahua_df, args.output)
