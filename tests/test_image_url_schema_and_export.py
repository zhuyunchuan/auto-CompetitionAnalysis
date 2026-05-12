from datetime import datetime

from openpyxl import load_workbook

from src.export.excel_writer import ExcelWriter
from src.storage.db import Database
from src.storage.schema import ProductCatalog, RunSummary


def test_product_catalog_schema_migration_adds_image_url(tmp_path):
    db_path = tmp_path / "test.db"
    db = Database(db_path=str(db_path))

    engine = db.engine
    with engine.begin() as conn:
        conn.exec_driver_sql(
            """
            CREATE TABLE product_catalog (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id VARCHAR(50) NOT NULL,
                brand VARCHAR(100) NOT NULL,
                series_l1 VARCHAR(200),
                series_l2 VARCHAR(200),
                product_model VARCHAR(200) NOT NULL,
                product_name VARCHAR(500) NOT NULL,
                product_url VARCHAR(1000) NOT NULL,
                locale VARCHAR(20) NOT NULL DEFAULT 'en-US',
                first_seen_at DATETIME NOT NULL,
                last_seen_at DATETIME NOT NULL,
                catalog_status VARCHAR(50) NOT NULL DEFAULT 'current'
            )
            """
        )

    db.init_db()

    with engine.begin() as conn:
        cols = conn.exec_driver_sql("PRAGMA table_info(product_catalog)").fetchall()
        col_names = {row[1] for row in cols}
        assert "image_url" in col_names


def test_excel_catalog_has_image_url_column(tmp_path):
    writer = ExcelWriter(output_dir=tmp_path)
    run_id = "test_run"

    catalog = ProductCatalog(
        run_id=run_id,
        brand="hikvision",
        series_l1="Pro",
        series_l2="Test",
        product_model="DS-TEST",
        product_name="DS-TEST",
        product_url="https://example.com",
        image_url="https://example.com/a.jpg",
        locale="en-US",
        first_seen_at=datetime.utcnow(),
        last_seen_at=datetime.utcnow(),
        catalog_status="current",
    )

    summary = RunSummary(
        run_id=run_id,
        schedule_type="manual",
        started_at=datetime.utcnow(),
        ended_at=datetime.utcnow(),
        catalog_count=1,
        spec_field_count=0,
        issue_count=0,
        new_series_count=0,
        disappeared_series_count=0,
        success_rate=1.0,
        status="completed",
    )

    out_path = writer.generate_report(
        run_id=run_id,
        catalog_data={"hikvision": [catalog], "dahua": []},
        spec_data={"hikvision": [], "dahua": []},
        issues=[],
        summary=summary,
    )

    wb = load_workbook(out_path)
    ws = wb["hikvision_catalog"]
    header = [cell.value for cell in ws[1]]
    assert "image_url" in header
