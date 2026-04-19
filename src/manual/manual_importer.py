"""
Manual input importer for competitor analysis system.

This module imports manual corrections and additions from Excel files,
validates the data, and stores it in the manual_inputs table for later
application to specification records.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import uuid

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from src.storage.schema import ManualInput
from src.core.logging import get_logger

logger = get_logger(__name__)


class ManualImporter:
    """
    Import manual inputs from Excel files.

    Reads the manual_append sheet from exported Excel files,
    validates required fields, and imports into manual_inputs table.
    """

    REQUIRED_FIELDS = [
        'brand',
        'series_l1',
        'series_l2',
        'product_model',
        'field_code',
        'manual_value',
        'operator',
        'reason',
    ]

    def __init__(self, session: Session):
        """
        Initialize manual importer.

        Args:
            session: SQLAlchemy session
        """
        self.session = session

    def import_from_excel(
        self,
        excel_path: Path,
        sheet_name: str = 'manual_append',
        run_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Import manual inputs from Excel file.

        Args:
            excel_path: Path to Excel file
            sheet_name: Name of sheet containing manual inputs
            run_id: Optional run ID to associate with imports

        Returns:
            Dictionary with import results:
                - success_count: Number of records imported
                - error_count: Number of records with errors
                - errors: List of error messages
        """
        logger.info(
            f"Importing manual inputs from Excel",
            extra={
                "excel_path": str(excel_path),
                "sheet_name": sheet_name,
            }
        )

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        ws = wb[sheet_name]

        # Read header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {col: idx for idx, col in enumerate(header_row) if col}

        # Validate required columns exist
        missing_cols = set(self.REQUIRED_FIELDS) - set(header_map.keys())
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")

        # Import data rows
        results = {
            'success_count': 0,
            'error_count': 0,
            'errors': [],
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            result = self._import_row(row, header_map, row_idx, run_id)

            if result['success']:
                results['success_count'] += 1
            else:
                results['error_count'] += 1
                results['errors'].append(result['error'])

        self.session.flush()

        logger.info(
            f"Manual input import completed",
            extra={
                "excel_path": str(excel_path),
                "success_count": results['success_count'],
                "error_count": results['error_count'],
            }
        )

        return results

    def _import_row(
        self,
        row: tuple,
        header_map: Dict[str, int],
        row_idx: int,
        run_id: Optional[str],
    ) -> Dict[str, Any]:
        """
        Import a single row from Excel.

        Args:
            row: Row data tuple
            header_map: Mapping of column names to indices
            row_idx: Row number (for error reporting)
            run_id: Optional run ID

        Returns:
            Dictionary with import result
        """
        try:
            # Extract values from row
            def get_value(col_name: str) -> Any:
                idx = header_map.get(col_name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            brand = get_value('brand')
            series_l1 = get_value('series_l1')
            series_l2 = get_value('series_l2')
            product_model = get_value('product_model')
            field_code = get_value('field_code')
            manual_value = get_value('manual_value')
            operator = get_value('operator')
            reason = get_value('reason')

            # Validate required fields
            if not all([brand, series_l1, series_l2, product_model, field_code, manual_value, operator, reason]):
                missing = [
                    field for field in self.REQUIRED_FIELDS
                    if not get_value(field)
                ]
                return {
                    'success': False,
                    'error': f"Row {row_idx}: Missing required fields: {missing}",
                }

            # Validate hierarchy completeness
            if not brand or not series_l1 or not series_l2:
                return {
                    'success': False,
                    'error': f"Row {row_idx}: Incomplete hierarchy (brand, series_l1, series_l2 required)",
                }

            # Check for duplicate
            existing = self.session.query(ManualInput).filter(
                ManualInput.brand == brand,
                ManualInput.series_l1 == series_l1,
                ManualInput.series_l2 == series_l2,
                ManualInput.product_model == product_model,
                ManualInput.field_code == field_code,
            ).first()

            if existing:
                # Update existing record
                existing.manual_value = str(manual_value)
                existing.operator = str(operator)
                existing.reason = str(reason)
                existing.created_at = datetime.utcnow()

                logger.debug(
                    f"Updated existing manual input",
                    extra={
                        "brand": brand,
                        "model": product_model,
                        "field_code": field_code,
                    }
                )
            else:
                # Create new record
                manual_input = ManualInput(
                    input_id=str(uuid.uuid4()),
                    brand=str(brand),
                    series_l1=str(series_l1),
                    series_l2=str(series_l2),
                    product_model=str(product_model),
                    field_code=str(field_code),
                    manual_value=str(manual_value),
                    operator=str(operator),
                    reason=str(reason),
                    created_at=datetime.utcnow(),
                )

                self.session.add(manual_input)

                logger.debug(
                    f"Created manual input",
                    extra={
                        "input_id": manual_input.input_id,
                        "brand": brand,
                        "model": product_model,
                        "field_code": field_code,
                    }
                )

            return {'success': True}

        except Exception as e:
            logger.exception(
                f"Error importing row {row_idx}",
                extra={"row_idx": row_idx}
            )
            return {
                'success': False,
                'error': f"Row {row_idx}: {str(e)}",
            }

    def get_pending_inputs(
        self,
        run_id: Optional[str] = None,
    ) -> List[ManualInput]:
        """
        Get pending manual inputs for application.

        Args:
            run_id: Optional run ID filter

        Returns:
            List of ManualInput instances
        """
        query = self.session.query(ManualInput).order_by(
            ManualInput.created_at.desc()
        )

        # Note: Manual inputs don't have run_id in schema
        # They are global and can be applied to any run
        # If filtering by run_id is needed, it would require a join table
        # or additional logic

        results = query.all()

        logger.debug(
            f"Retrieved pending manual inputs",
            extra={"count": len(results)}
        )

        return results

    def delete_input(self, input_id: str) -> bool:
        """
        Delete a manual input by ID.

        Args:
            input_id: Input ID to delete

        Returns:
            True if deleted, False if not found
        """
        count = self.session.query(ManualInput).filter(
            ManualInput.input_id == input_id
        ).delete()

        if count > 0:
            logger.info(
                f"Deleted manual input",
                extra={"input_id": input_id}
            )
            return True

        return False

    def get_inputs_for_product(
        self,
        brand: str,
        product_model: str,
    ) -> List[ManualInput]:
        """
        Get all manual inputs for a specific product.

        Args:
            brand: Brand name
            product_model: Product model

        Returns:
            List of ManualInput instances
        """
        results = self.session.query(ManualInput).filter(
            ManualInput.brand == brand,
            ManualInput.product_model == product_model,
        ).all()

        return results

    def get_inputs_for_series(
        self,
        brand: str,
        series_l1: str,
        series_l2: Optional[str] = None,
    ) -> List[ManualInput]:
        """
        Get all manual inputs for a series.

        Args:
            brand: Brand name
            series_l1: Series level 1
            series_l2: Optional series level 2

        Returns:
            List of ManualInput instances
        """
        conditions = [
            ManualInput.brand == brand,
            ManualInput.series_l1 == series_l1,
        ]

        if series_l2:
            conditions.append(ManualInput.series_l2 == series_l2)

        results = self.session.query(ManualInput).filter(
            *conditions
        ).all()

        return results
