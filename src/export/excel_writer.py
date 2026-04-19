"""
Excel report writer for competitor analysis system.

This module generates comprehensive Excel reports with multiple sheets including
catalog data, specifications, quality issues, and manual input templates.
Uses openpyxl for advanced formatting and conditional formatting.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from src.storage.schema import ProductCatalog, ProductSpecLong, DataQualityIssue, RunSummary
from src.core.logging import get_logger

logger = get_logger(__name__)


class ExcelWriter:
    """
    Generate Excel reports from database data.

    Creates multi-sheet Excel workbooks with:
    - Brand-specific catalog and specification sheets
    - Manual input template sheet
    - Data quality issues sheet with conditional formatting
    - Run summary with metrics
    """

    # Sheet names (fixed order)
    SHEET_CATALOGS = {
        'hikvision': 'hikvision_catalog',
        'dahua': 'dahua_catalog',
    }
    SHEET_SPECS = {
        'hikvision': 'hikvision_specs',
        'dahua': 'dahua_specs',
    }
    SHEET_MANUAL_APPEND = 'manual_append'
    SHEET_QUALITY_ISSUES = 'data_quality_issues'
    SHEET_RUN_SUMMARY = 'run_summary'

    # Column definitions (fixed order per specification)
    CATALOG_COLUMNS = [
        'brand',
        'series_l1',
        'series_l2',
        'product_model',
        'product_name',
        'product_url',
        'catalog_status',
    ]

    SPEC_COLUMNS = [
        'brand',
        'series_l1',
        'series_l2',
        'product_model',
        'field_code',
        'field_name',
        'raw_value',
        'normalized_value',
        'unit',
        'extract_confidence',
        'is_manual_override',
    ]

    MANUAL_APPEND_COLUMNS = [
        'brand',
        'series_l1',
        'series_l2',
        'product_model',
        'field_code',
        'manual_value',
        'operator',
        'reason',
    ]

    ISSUE_COLUMNS = [
        'run_id',
        'brand',
        'series_l1',
        'series_l2',
        'product_model',
        'issue_type',
        'field_code',
        'issue_detail',
        'severity',
        'status',
    ]

    RUN_SUMMARY_COLUMNS = [
        'run_id',
        'schedule_type',
        'started_at',
        'ended_at',
        'catalog_count',
        'spec_field_count',
        'issue_count',
        'new_series_count',
        'disappeared_series_count',
        'success_rate',
        'status',
    ]

    # Formatting styles
    HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Severity colors
    SEVERITY_COLORS = {
        'P1': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'P2': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        'P3': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    }

    def __init__(self, output_dir: Path):
        """
        Initialize Excel writer.

        Args:
            output_dir: Directory to write Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        run_id: str,
        catalog_data: Dict[str, List[ProductCatalog]],
        spec_data: Dict[str, List[ProductSpecLong]],
        issues: List[DataQualityIssue],
        summary: RunSummary,
    ) -> Path:
        """
        Generate complete Excel report for a run.

        Args:
            run_id: Run identifier
            catalog_data: Dictionary mapping brand to catalog entries
            spec_data: Dictionary mapping brand to spec records
            issues: List of quality issues
            summary: Run summary record

        Returns:
            Path to generated Excel file
        """
        filename = f"competitor_specs_{run_id}.xlsx"
        filepath = self.output_dir / filename

        logger.info(
            f"Generating Excel report",
            extra={
                "run_id": run_id,
                "filepath": str(filepath),
            }
        )

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets in fixed order
        self._create_catalog_sheets(wb, catalog_data)
        self._create_spec_sheets(wb, spec_data)
        self._create_manual_append_sheet(wb)
        self._create_quality_issues_sheet(wb, issues)
        self._create_run_summary_sheet(wb, summary)

        # Save workbook
        wb.save(filepath)

        logger.info(
            f"Excel report generated successfully",
            extra={
                "run_id": run_id,
                "filepath": str(filepath),
                "file_size_kb": filepath.stat().st_size / 1024,
            }
        )

        return filepath

    def _create_catalog_sheets(
        self,
        wb: Workbook,
        catalog_data: Dict[str, List[ProductCatalog]],
    ) -> None:
        """
        Create catalog sheets for each brand.

        Args:
            wb: Workbook instance
            catalog_data: Dictionary mapping brand to catalog entries
        """
        for brand, sheet_name in self.SHEET_CATALOGS.items():
            if brand not in catalog_data:
                continue

            ws = wb.create_sheet(title=sheet_name)
            records = catalog_data[brand]

            # Write header
            for col_idx, col_name in enumerate(self.CATALOG_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.BORDER_THIN

            # Write data rows
            for row_idx, record in enumerate(records, 2):
                ws.cell(row=row_idx, column=1, value=record.brand)
                ws.cell(row=row_idx, column=2, value=record.series_l1)
                ws.cell(row=row_idx, column=3, value=record.series_l2)
                ws.cell(row=row_idx, column=4, value=record.product_model)
                ws.cell(row=row_idx, column=5, value=record.product_name)
                ws.cell(row=row_idx, column=6, value=record.product_url)
                ws.cell(row=row_idx, column=7, value=record.catalog_status)

            # Apply borders to all data cells
            self._apply_borders(ws, max_row=len(records) + 1, max_col=len(self.CATALOG_COLUMNS))

            # Auto-adjust column widths
            self._auto_adjust_columns(ws, self.CATALOG_COLUMNS)

            # Freeze header row
            ws.freeze_panes = 'A2'

            logger.debug(
                f"Created catalog sheet for {brand}",
                extra={
                    "sheet": sheet_name,
                    "row_count": len(records),
                }
            )

    def _create_spec_sheets(
        self,
        wb: Workbook,
        spec_data: Dict[str, List[ProductSpecLong]],
    ) -> None:
        """
        Create specification sheets for each brand.

        Args:
            wb: Workbook instance
            spec_data: Dictionary mapping brand to spec records
        """
        for brand, sheet_name in self.SHEET_SPECS.items():
            if brand not in spec_data:
                continue

            ws = wb.create_sheet(title=sheet_name)
            records = spec_data[brand]

            # Write header
            for col_idx, col_name in enumerate(self.SPEC_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.BORDER_THIN

            # Write data rows
            for row_idx, record in enumerate(records, 2):
                ws.cell(row=row_idx, column=1, value=record.brand)
                ws.cell(row=row_idx, column=2, value=record.series_l1)
                ws.cell(row=row_idx, column=3, value=record.series_l2)
                ws.cell(row=row_idx, column=4, value=record.product_model)
                ws.cell(row=row_idx, column=5, value=record.field_code)
                ws.cell(row=row_idx, column=6, value=record.field_name)
                ws.cell(row=row_idx, column=7, value=record.raw_value)
                ws.cell(row=row_idx, column=8, value=record.normalized_value)
                ws.cell(row=row_idx, column=9, value=record.unit)
                ws.cell(row=row_idx, column=10, value=record.extract_confidence)
                ws.cell(row=row_idx, column=11, value=record.is_manual_override)

            # Apply borders to all data cells
            self._apply_borders(ws, max_row=len(records) + 1, max_col=len(self.SPEC_COLUMNS))

            # Auto-adjust column widths
            self._auto_adjust_columns(ws, self.SPEC_COLUMNS)

            # Freeze header row
            ws.freeze_panes = 'A2'

            logger.debug(
                f"Created spec sheet for {brand}",
                extra={
                    "sheet": sheet_name,
                    "row_count": len(records),
                }
            )

    def _create_manual_append_sheet(self, wb: Workbook) -> None:
        """
        Create manual input template sheet.

        This sheet provides a template for users to submit manual corrections
        or additions. It contains only headers with empty rows for data entry.

        Args:
            wb: Workbook instance
        """
        ws = wb.create_sheet(title=self.SHEET_MANUAL_APPEND)

        # Write header
        for col_idx, col_name in enumerate(self.MANUAL_APPEND_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Add 10 empty rows for data entry
        for row_idx in range(2, 12):
            for col_idx in range(1, len(self.MANUAL_APPEND_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.BORDER_THIN

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, self.MANUAL_APPEND_COLUMNS)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add comment/instruction in first cell
        ws['A12'] = 'Instructions: Fill in the rows above with manual corrections. Required fields: brand, series_l1, series_l2, product_model, field_code, manual_value, operator, reason.'
        ws['A12'].font = Font(italic=True, color='666666', size=9)

        logger.debug(f"Created manual append template sheet")

    def _create_quality_issues_sheet(
        self,
        wb: Workbook,
        issues: List[DataQualityIssue],
    ) -> None:
        """
        Create data quality issues sheet with conditional formatting.

        Args:
            wb: Workbook instance
            issues: List of quality issues
        """
        ws = wb.create_sheet(title=self.SHEET_QUALITY_ISSUES)

        # Write header
        for col_idx, col_name in enumerate(self.ISSUE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Write data rows
        for row_idx, issue in enumerate(issues, 2):
            ws.cell(row=row_idx, column=1, value=issue.run_id)
            ws.cell(row=row_idx, column=2, value=issue.brand)
            ws.cell(row=row_idx, column=3, value=issue.series_l1)
            ws.cell(row=row_idx, column=4, value=issue.series_l2)
            ws.cell(row=row_idx, column=5, value=issue.product_model)
            ws.cell(row=row_idx, column=6, value=issue.issue_type)
            ws.cell(row=row_idx, column=7, value=issue.field_code)
            ws.cell(row=row_idx, column=8, value=issue.issue_detail)
            cell_severity = ws.cell(row=row_idx, column=9, value=issue.severity)
            ws.cell(row=row_idx, column=10, value=issue.status)

            # Apply severity-based color coding
            if issue.severity in self.SEVERITY_COLORS:
                cell_severity.fill = self.SEVERITY_COLORS[issue.severity]

        # Apply borders to all data cells
        self._apply_borders(ws, max_row=len(issues) + 1, max_col=len(self.ISSUE_COLUMNS))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, self.ISSUE_COLUMNS)

        # Freeze header row
        ws.freeze_panes = 'A2'

        logger.debug(
            f"Created quality issues sheet",
            extra={"row_count": len(issues)}
        )

    def _create_run_summary_sheet(
        self,
        wb: Workbook,
        summary: RunSummary,
    ) -> None:
        """
        Create run summary sheet with metrics.

        Args:
            wb: Workbook instance
            summary: Run summary record
        """
        ws = wb.create_sheet(title=self.SHEET_RUN_SUMMARY)

        # Write header
        for col_idx, col_name in enumerate(self.RUN_SUMMARY_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Write data row
        ws.cell(row=2, column=1, value=summary.run_id)
        ws.cell(row=2, column=2, value=summary.schedule_type)
        ws.cell(row=2, column=3, value=summary.started_at.strftime('%Y-%m-%d %H:%M:%S') if summary.started_at else '')
        ws.cell(row=2, column=4, value=summary.ended_at.strftime('%Y-%m-%d %H:%M:%S') if summary.ended_at else '')
        ws.cell(row=2, column=5, value=summary.catalog_count)
        ws.cell(row=2, column=6, value=summary.spec_field_count)
        ws.cell(row=2, column=7, value=summary.issue_count)
        ws.cell(row=2, column=8, value=summary.new_series_count)
        ws.cell(row=2, column=9, value=summary.disappeared_series_count)
        ws.cell(row=2, column=10, value=f"{summary.success_rate:.2%}")
        ws.cell(row=2, column=11, value=summary.status)

        # Apply borders
        self._apply_borders(ws, max_row=2, max_col=len(self.RUN_SUMMARY_COLUMNS))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, self.RUN_SUMMARY_COLUMNS)

        # Freeze header row
        ws.freeze_panes = 'A2'

        logger.debug(f"Created run summary sheet")

    def _apply_borders(
        self,
        ws,
        max_row: int,
        max_col: int,
    ) -> None:
        """
        Apply borders to a range of cells.

        Args:
            ws: Worksheet instance
            max_row: Maximum row number
            max_col: Maximum column number
        """
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER_THIN

    def _auto_adjust_columns(
        self,
        ws,
        columns: List[str],
        min_width: int = 12,
        max_width: int = 50,
    ) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet instance
            columns: List of column names
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for col_idx, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)

            # Calculate width based on header and data
            width = len(str(col_name)) + 2  # Header width + padding

            # Check first 100 data rows for content width
            for row in range(2, min(102, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    # Account for newlines in wrapped text
                    cell_length = max(
                        cell_length,
                        max(len(line) for line in str(cell_value).split('\n'))
                    )
                    width = max(width, cell_length + 2)

            # Clamp to min/max bounds
            width = max(min_width, min(width, max_width))

            ws.column_dimensions[col_letter].width = width
